import openpyxl

# Make a new file
xlFile = openpyxl.Workbook()
xlFile.save("testFile.xlsx")

# You need to declare a sheet to work on
page1 = xlFile.active

page1['C2'] = "This Cell is C2"

# Renaming Worksheet
page1.title = "firstsheet"

# Make Another Page
xlFile.create_sheet("secondSheet")

# Must save at the end
xlFile.save("testFile.xlsx")

# Reading Workbook
xlFile = openpyxl.load_workbook("testFile.xlsx")
currentSheet = xlFile.active

# Print rows and columns
print("Rows: " + str(currentSheet.max_row))
print("Columns: " + str(currentSheet.max_column))

# Cell Indexing
print("Value in Call C2 is " + currentSheet['C2'].value)

# Reading multiple Cells
# Example: Print all the headers
values = [currentSheet.cell(row = 1, column = 1).value for i in range(1, currentSheet.max_column + 1)]
print(values)

# Printing first 10 rows from a range of columns
myList = list()

for value in currentSheet.iter_rows(min_row = 1, max_row = 11, min_col = 1, max_col = 6, values_only = True):
    myList.append(value)

for i in myList:
    print(i)