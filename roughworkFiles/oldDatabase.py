import openpyxl
import pandas as pd

def sortDatabase(dataframe):
    return dataframe.sort_values("Card No.")
    
def removeDuplicates(dataframe):
    return dataframe.drop_duplicates("Card No.")

def convertToPanda():
    return pd.read_excel("cfvdatabase.xlsx")

def filterDatabase(keyword):
    path = r"C:\Users\ryanh\Documents\Personal\Coding\carddatabase\cfvDatabase.xlsx"

    cfvDataframe = convertToPanda()
    values = (cfvDataframe[keyword].unique())
    values.sort()

    for word in values:
        filteredDataframe = cfvDataframe[cfvDataframe[keyword] == word]
        del filteredDataframe[keyword]

        with pd.ExcelWriter(path, mode = "a", engine = "openpyxl") as writer:
            filteredDataframe.to_excel(writer, sheet_name = word, index = False)

def formatDatabase():
    cfvDataframe = convertToPanda()
    cfvDataframe = removeDuplicates(cfvDataframe)
    cfvDataframe = sortDatabase(cfvDataframe)
    cfvDataframe.to_excel("cfvdatabase.xlsx", sheet_name = "All Cards", index = False)

def addHeaders(page):
    headers = ["Card No.", "Name", "Card Type", "Grade", "Skill", "Imaginary Gift", "Special Icon", 
               "Trigger Effect", "Power", "Shield", "Critical", "Nation", "Clan", "Race", "Format", 
               "Artist", "Full Art(s)", "Card Set(s)", "Rarity", "Card Effect(s)", "Release Date"]

    index = 0
    for keyword in headers:
        page.cell(row = 1, column = index + 1).value = keyword
        index = index + 1

def createPage(spreadsheet, name):
    newPage = spreadsheet.create_sheet(name)
    addHeaders(newPage)

    return newPage

def createDatabase():
    spreadsheet = openpyxl.Workbook()
    currentPage = spreadsheet.active
    currentPage.title = "All Cards"

    addHeaders(currentPage)

    spreadsheet.save("cfvdatabase.xlsx")

def writeCardInfo(cardDictionary):
    spreadsheet = openpyxl.load_workbook("cfvdatabase.xlsx")
    currentPage = spreadsheet.active

    headers = []    
    for index in range(1, currentPage.max_column + 1):
        headers.append(currentPage.cell(row = 1, column = index).value)

    cardData = []
    for keyword in headers:
        if (cardDictionary.get(keyword) == None):
            cardData.append("-")
        else:
            cardData.append(str(cardDictionary.get(keyword)))
    
    currentPage.append(tuple(cardData))
    spreadsheet.save("cfvdatabase.xlsx")